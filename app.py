import streamlit as st
import pandas as pd
import requests
import folium
from streamlit_folium import st_folium
from ortools.constraint_solver import routing_enums_pb2
from ortools.constraint_solver import pywrapcp
from haversine import haversine, Unit
import io
import datetime
import re
import time 
import urllib.parse
import openpyxl 

# --- CONFIGURACIÓN DE LA PÁGINA ---
st.set_page_config(page_title="Gestor de Rutas Logísticas", layout="wide")
st.title("🚚 Gestor de Rutas y Capas")

# --- API KEY PREDETERMINADA Y PERSONALIZADA ---
api_key_default = "5b3ce3597851110001cf62480080f4189d6143db946e7c7267b9343d"

st.sidebar.header("🔑 Conexión OpenRouteService")
api_key_user = st.sidebar.text_input("API Key propia (Solo si te quedas sin saldo diario)", type="password", help="Si te sale el error 'Quota exceeded', pega aquí tu propia API Key.")
api_key = api_key_user if api_key_user else api_key_default

if 'calculo_terminado' not in st.session_state:
    st.session_state['calculo_terminado'] = False

# --- SCRIPT DE BOTONES FLOTANTES PARA EL MAPA ---
js_toggle_capas = """
<div id="panel-botones-capas" style="position: absolute; top: 12px; left: 55px; z-index: 9999; background: white; padding: 6px; border-radius: 5px; box-shadow: 0 1px 5px rgba(0,0,0,0.65); font-family: sans-serif; font-size: 13px;">
    <button type="button" onclick="toggleFoliumLayers(false, event)" style="cursor: pointer; background: #ffebee; border: 1px solid #ffcdd2; padding: 4px 8px; border-radius: 3px; font-weight: bold; margin-right: 5px; color: #b71c1c;">❌ Apagar Todo</button>
    <button type="button" onclick="toggleFoliumLayers(true, event)" style="cursor: pointer; background: #e8f5e9; border: 1px solid #c8e6c9; padding: 4px 8px; border-radius: 3px; font-weight: bold; color: #1b5e20;">✅ Prender Todo</button>
</div>
<script>
function toggleFoliumLayers(turnOn, e) {
    if(e) {
        e.stopPropagation();
        e.preventDefault();
    }
    var checkboxes = document.querySelectorAll('.leaflet-control-layers-selector');
    checkboxes.forEach(function(cb) {
        if((turnOn && !cb.checked) || (!turnOn && cb.checked)) {
            cb.click();
        }
    });
}

var panel = document.getElementById('panel-botones-capas');
if(panel) {
    panel.addEventListener('mousedown', function(e){ e.stopPropagation(); });
    panel.addEventListener('dblclick', function(e){ e.stopPropagation(); });
    panel.addEventListener('wheel', function(e){ e.stopPropagation(); });
}
</script>
"""

# --- BARRA LATERAL: CARGA ---
st.sidebar.markdown("---")
st.sidebar.header("Carga de Datos")
archivo_subido = st.sidebar.file_uploader("Sube tu archivo Excel", type=["xlsx", "xls"])

# --- FUNCIONES AUXILIARES ---
def preparar_coordenadas(coord_str):
    try:
        partes = str(coord_str).split(',')
        if len(partes) >= 2:
            lat = float(partes[0].strip())
            lon = float(partes[1].strip())
            return [lon, lat]
        return None
    except Exception:
        return None

def limpiar_nombre_excel(nombre):
    return re.sub(r'[\\/*?:\[\]]', '_', str(nombre))[:31]

def dibujar_geozona_circular(coordenadas_lon_lat, nombre_capa, color, mapa, mostrar_por_defecto=True):
    coords_lat_lon = [(p[1], p[0]) for p in coordenadas_lon_lat]
    if len(coords_lat_lon) > 0:
        avg_lat = sum([p[0] for p in coords_lat_lon]) / len(coords_lat_lon)
        avg_lon = sum([p[1] for p in coords_lat_lon]) / len(coords_lat_lon)
        centro = (avg_lat, avg_lon)

        max_radio_metros = 0
        for punto in coords_lat_lon:
            dist = haversine(centro, punto, unit=Unit.METERS)
            if dist > max_radio_metros:
                max_radio_metros = dist
        
        radio_final = max_radio_metros * 1.05 if max_radio_metros > 10 else 200

        capa = folium.FeatureGroup(name=nombre_capa, show=mostrar_por_defecto)
        folium.Circle(
            location=centro, radius=radio_final, color=color,
            fill=True, fill_color=color, fill_opacity=0.15, weight=2
        ).add_to(capa)
        capa.add_to(mapa)

# --- CONEXIÓN PURA BLINDADA CONTRA CAÍDAS ---
def pedir_matriz_ors_con_reintento(body, headers):
    for intento in range(5): 
        try:
            resp = requests.post('https://api.openrouteservice.org/v2/matrix/driving-car', json=body, headers=headers, timeout=70)
            if resp.status_code == 200:
                return resp.json(), None
            elif "Quota" in resp.text or resp.status_code == 403:
                return None, "QUOTA_EXCEEDED"
            elif resp.status_code == 429 or "Rate limit" in resp.text:
                time.sleep(60) 
            elif resp.status_code >= 500 or "unknown" in resp.text.lower():
                time.sleep(5)
            else:
                return None, resp.text
        except requests.exceptions.RequestException:
            time.sleep(5)
    return None, "Superado el límite de reintentos. El servidor está caído temporalmente."

def pedir_trazado_ors_con_reintento(body, headers):
    for intento in range(5):
        try:
            resp = requests.post('https://api.openrouteservice.org/v2/directions/driving-car/geojson', json=body, headers=headers, timeout=70)
            if resp.status_code == 200:
                return resp.json(), None
            elif "Quota" in resp.text or resp.status_code == 403:
                return None, "QUOTA_EXCEEDED"
            elif resp.status_code == 429 or "Rate limit" in resp.text:
                time.sleep(60)
            elif resp.status_code >= 500 or "unknown" in resp.text.lower():
                time.sleep(5)
            else:
                return None, resp.text
        except requests.exceptions.RequestException:
            time.sleep(5)
    return None, "Superado el límite de reintentos."

def obtener_matriz_masiva(lista_coords, headers):
    N = len(lista_coords)
    if N <= 50:
        body_matrix = {"locations": lista_coords, "metrics": ["distance", "duration"]}
        data, err = pedir_matriz_ors_con_reintento(body_matrix, headers)
        if data:
            return data['distances'], data['durations'], None
        return None, None, err
    else:
        matriz_dist = [[0.0] * N for _ in range(N)]
        matriz_dur = [[0.0] * N for _ in range(N)]
        chunk_size = 15 
        for i in range(0, N, chunk_size):
            for j in range(0, N, chunk_size):
                src_chunk = lista_coords[i : i+chunk_size]
                dst_chunk = lista_coords[j : j+chunk_size]
                locs = []
                src_indices, dst_indices = [], []
                for pt in src_chunk:
                    if pt not in locs: locs.append(pt)
                    src_indices.append(locs.index(pt))
                for pt in dst_chunk:
                    if pt not in locs: locs.append(pt)
                    dst_indices.append(locs.index(pt))
                body_matrix = {"locations": locs, "sources": src_indices, "destinations": dst_indices, "metrics": ["distance", "duration"]}
                data, err = pedir_matriz_ors_con_reintento(body_matrix, headers)
                if data:
                    dists, durs = data['distances'], data['durations']
                    for u in range(len(src_chunk)):
                        for v in range(len(dst_chunk)):
                            matriz_dist[i+u][j+v] = dists[u][v]
                            matriz_dur[i+u][j+v] = durs[u][v]
                else:
                    return None, None, err
                time.sleep(1.2) 
        return matriz_dist, matriz_dur, None

def obtener_trazado_masivo(coords_ordenadas, headers):
    total_dist, total_dur = 0, 0
    all_segments, merged_coordinates = [], []
    chunk_size = 40
    for i in range(0, len(coords_ordenadas) - 1, chunk_size - 1):
        chunk = coords_ordenadas[i:i + chunk_size]
        if len(chunk) < 2: break
        body_dirs = {"coordinates": chunk, "radiuses": [-1] * len(chunk)}
        data, err = pedir_trazado_ors_con_reintento(body_dirs, headers)
        if data:
            props = data['features'][0]['properties']['summary']
            segs = data['features'][0]['properties'].get('segments', [])
            geom = data['features'][0]['geometry']['coordinates']
            total_dist += props['distance']
            total_dur += props['duration']
            all_segments.extend(segs)
            if not merged_coordinates:
                merged_coordinates.extend(geom)
            else:
                merged_coordinates.extend(geom[1:])
        else:
            return None, err
        time.sleep(1.5)
    fake_geojson = {
        "type": "FeatureCollection",
        "features": [{
            "type": "Feature",
            "properties": {"summary": {"distance": total_dist, "duration": total_dur}, "segments": all_segments},
            "geometry": {"type": "LineString", "coordinates": merged_coordinates}
        }]
    }
    return fake_geojson, None

# =====================================================================
# BLOQUE DE SEGURIDAD ABSOLUTA: Nada se ejecuta si no hay archivo
# =====================================================================
if archivo_subido is None:
    st.info("👈 Por favor, sube tu archivo Excel en la barra lateral para comenzar.")
else:
    # --- PROCESAMIENTO INICIAL DEL DATAFRAME ---
    df = pd.read_excel(archivo_subido)
    df.columns = df.columns.str.strip() 

    # -------------------------------------------------------------
    # IDENTIFICADOR INTELIGENTE DE ARCHIVO DESCARGADO
    # -------------------------------------------------------------
    columnas_requeridas_cronograma = ['Orden', 'Día', 'Ruta', 'Departamento', 'Lugar', 'Coordenadas', 'Llegada', 'Salida', 'Minutos Tramo', 'Minutos Acumulados', 'Km Tramo', 'Km Acumulados']
    es_cronograma_descargado = all(col in df.columns for col in columnas_requeridas_cronograma)

    if 'Ruta' not in df.columns:
        df['Ruta'] = "Sin Asignar"
    df['Ruta'] = df['Ruta'].fillna("Sin Asignar")

    df['Coords_Procesadas'] = df['Coordenadas'].apply(preparar_coordenadas)
    df = df.dropna(subset=['Coords_Procesadas']).copy()

    if df.empty:
        st.error("❌ No se encontraron coordenadas válidas en el archivo.")
        st.stop()

    # ======================================================================
    # FLUJO 1: ARCHIVO YA CALCULADO (MODO VISUALIZACIÓN RÁPIDA)
    # ======================================================================
    if es_cronograma_descargado:
        tipo_ruteo = "Visualización Descargada"
        st.success("📄 ¡Cronograma Detallado detectado! Se ha evitado el menú de configuración.")
        st.info("Presiona el botón para procesar el recorrido guardado en tu archivo Excel.")
        
        if st.button("🗺️ Mostrar Ruteo", type="primary", use_container_width=True):
            with st.spinner("Procesando trazados desde el archivo..."):
                df_valido = df.sort_values(by=['Día', 'Ruta', 'Orden'])
                
                colores = ['blue', 'red', 'green', 'purple', 'orange', 'darkred', 'cadetblue', 'darkblue', 'pink', 'lightgreen']
                datos_para_resumen = []
                color_idx = 0
                headers = {'Authorization': api_key, 'Content-Type': 'application/json'}
                
                grupos = df_valido.groupby(['Día', 'Ruta'], sort=False)
                
                for (dia, ruta), df_grupo in grupos:
                    id_unico = f"{dia} - {ruta}"
                    color_actual = colores[color_idx % len(colores)]
                    color_idx += 1
                    
                    coords_ordenadas = df_grupo['Coords_Procesadas'].tolist()
                    
                    if len(coords_ordenadas) > 1:
                        geojson, err_dirs = obtener_trazado_masivo(coords_ordenadas, headers)
                        if err_dirs:
                            st.error(f"Error trazando {ruta}: {err_dirs}")
                            continue
                            
                        props = geojson['features'][0]['properties']['summary']
                        segments = geojson['features'][0]['properties'].get('segments', [])
                        
                        paradas_info = []
                        for _, row in df_grupo.iterrows():
                            paradas_info.append({
                                "Día": row.get('Día',''), "Ruta": row.get('Ruta',''),
                                "Departamento": row.get('Departamento',''), "Lugar": row.get('Lugar',''),
                                "Coordenadas": row.get('Coordenadas',''), "Orden": row.get('Orden', '')
                            })
                            
                        datos_para_resumen.append({
                            "id_unico": id_unico, "dia": dia, "ruta": ruta,
                            "puntos": len(df_grupo), "dist_km": round(props['distance']/1000, 2),
                            "drive_mins": round(props['duration']/60, 0), "color": color_actual,
                            "paradas": paradas_info, "segmentos": segments,
                            "geojson": geojson,
                            "coords_ordenadas": coords_ordenadas
                        })

                st.session_state['datos_resumen'] = datos_para_resumen
                st.session_state['calculo_terminado'] = True

    # ======================================================================
    # FLUJO 2: ARCHIVO CRUDO (LÓGICA ORIGINAL + NUEVA OPCIÓN v2)
    # ======================================================================
    else:
        punto_final_vrp = ""
        hora_salida_vrp = datetime.time(8, 0)
        hora_llegada_vrp = datetime.time(14, 30)
        min_parada_vrp = 15
        min_parada_global = 15
        
        st.sidebar.header("1. Filtro de Días")
        dias_disponibles = df['Día'].unique().tolist()
        todos_dias = st.sidebar.checkbox("✔️ Todos los Días", value=True)
        if todos_dias:
            dias_seleccionados = dias_disponibles
        else:
            dias_seleccionados = st.sidebar.multiselect("Días:", dias_disponibles, default=[])

        if not dias_seleccionados:
            st.sidebar.warning("Selecciona al menos un Día.")
            st.stop()

        df_filtrado_dias = df[df['Día'].isin(dias_seleccionados)]

        st.sidebar.markdown("---")
        st.sidebar.header("2. Estrategia de Ruteo")

        tipo_ruteo = st.sidebar.radio(
            "Selecciona cómo armar las rutas:",
            [
                "Ruteo según Excel (Orden Original)", 
                "Ruteo Optimizado (IA)", 
                "Ruteo Optimizado (IA) v2",
                "Creación de rutas propias (Ideal Libre)",
                "Creación de rutas propias (Departamental Flexible)",
                "Creación de rutas propias (Departamental Fijo)",
                "Creación de rutas propias (Ideal Libre - Patrón Fijo)",
                "Creación de rutas propias (Departamental Flexible - Patrón Fijo)",
                "Creación de rutas propias (Departamental Fijo - Patrón Fijo)"
            ]
        )

        usar_config_global_v2 = False
        opciones_inicio_global = {}
        opciones_deptos_dict_global = {}
        
        opciones_inicio_dict = {}
        opciones_anteante_dict = {} 
        opciones_antepenultimo_dict = {}
        opciones_penultimo_dict = {}
        opciones_fin_dict = {}
        opciones_deptos_dict = {}
        hora_salida_rutas_dict = {}  
        rutas_seleccionadas = []

        if tipo_ruteo in ["Ruteo según Excel (Orden Original)", "Ruteo Optimizado (IA)", "Ruteo Optimizado (IA) v2"]:
            st.sidebar.markdown("**Filtro de Rutas**")
            rutas_disponibles = df_filtrado_dias['Ruta'].unique().tolist()
            todas_rutas = st.sidebar.checkbox("✔️ Todas las Rutas", value=True)
            if todas_rutas:
                rutas_seleccionadas = rutas_disponibles
            else:
                rutas_seleccionadas = st.sidebar.multiselect("Rutas:", rutas_disponibles, default=[])

            if not rutas_seleccionadas:
                st.sidebar.warning("Selecciona al menos una ruta.")
                st.stop()
                
            st.sidebar.markdown("---")
            st.sidebar.markdown("**📍 Configuración de Ruta**")
            
            min_parada_global = st.sidebar.number_input("⏳ Minutos demora por parada:", min_value=0, value=15, step=1)
            
            if tipo_ruteo == "Ruteo Optimizado (IA)":
                st.sidebar.info("Elige la hora de salida. Puedes forzar el orden global de los últimos 4 puntos.")
            elif tipo_ruteo == "Ruteo Optimizado (IA) v2":
                st.sidebar.info("Elige la hora de salida y el orden de los últimos puntos ESPECÍFICO por cada Departamento. Si existe LABNU en la ruta, se usará automáticamente como el cierre final sin que tengas que seleccionarlo.")
                usar_config_global_v2 = st.sidebar.checkbox("✔️ Usar la misma configuración de cierre para TODOS los días", value=True)
            else:
                st.sidebar.info("Elige la hora de salida para cada ruta.")

            # --- LÓGICA V2 GLOBAL (LIMPIA DE LABNU EN PANTALLA) ---
            if tipo_ruteo == "Ruteo Optimizado (IA) v2" and usar_config_global_v2:
                st.sidebar.markdown("**⚙️ Configuración Global (Aplica a todos los días)**")
                for ruta in rutas_seleccionadas:
                    df_ruta_global = df_filtrado_dias[df_filtrado_dias['Ruta'] == ruta].reset_index(drop=True)
                    if df_ruta_global.empty: continue
                    st.sidebar.markdown(f"**Ruta:** {ruta}")
                    
                    # Filtramos LABNU del Inicio Global
                    lugares_lista_g = [loc for loc in df_ruta_global['Lugar'].unique().tolist() if str(df_ruta_global[df_ruta_global['Lugar']==loc]['Departamento'].iloc[0]).strip().upper() != 'LABNU']
                    opciones_lugar_g = ["🤖 IA Decide"] + lugares_lista_g
                    
                    sel_ini_g = st.sidebar.selectbox("Inicio Global:", opciones_lugar_g, index=0, key=f"ini_g_{ruta}")
                    opciones_inicio_global[ruta] = sel_ini_g
                    
                    deptos_lista_g = df_ruta_global['Departamento'].unique().tolist()
                    opciones_deptos_dict_global[ruta] = {}
                    
                    for dept in deptos_lista_g:
                        if pd.isna(dept) or str(dept).strip() == '': continue
                        dept_str = str(dept).strip()
                        
                        # IGNORAMOS LABNU EN LA INTERFAZ VISUAL COMPLETAMENTE
                        if dept_str.upper() == 'LABNU': 
                            continue
                        
                        st.sidebar.markdown(f"🔹 *Depto: {dept_str}*")
                        l_dept_g = df_ruta_global[df_ruta_global['Departamento'] == dept]['Lugar'].unique().tolist()
                        
                        # Por seguridad, si algún lugar se coló con nombre LABNU, lo volamos de la lista
                        l_dept_g = [loc for loc in l_dept_g if str(loc).strip().upper() != 'LABNU']
                        opc_dept_g = ["🤖 IA Decide"] + l_dept_g
                        
                        sel_aa_g = st.sidebar.selectbox("Ante-antepenúltimo:", opc_dept_g, index=0, key=f"aa_g_{ruta}_{dept_str}")
                        sel_a_g = st.sidebar.selectbox("Antepenúltimo:", opc_dept_g, index=0, key=f"a_g_{ruta}_{dept_str}")
                        sel_p_g = st.sidebar.selectbox("Penúltimo:", opc_dept_g, index=0, key=f"p_g_{ruta}_{dept_str}")
                        sel_f_g = st.sidebar.selectbox("Último:", opc_dept_g, index=0, key=f"f_g_{ruta}_{dept_str}")
                        
                        opciones_deptos_dict_global[ruta][dept_str] = {
                            'aa': sel_aa_g, 'a': sel_a_g, 'p': sel_p_g, 'f': sel_f_g
                        }
                st.sidebar.markdown("---")
                st.sidebar.markdown("**⏱️ Horarios de Salida por Día**")

            # --- LÓGICA DÍA POR DÍA O ASIGNACIÓN DE HORARIOS ---
            for dia in dias_seleccionados:
                for ruta in rutas_seleccionadas:
                    df_unicaruta = df_filtrado_dias[(df_filtrado_dias['Día'] == dia) & (df_filtrado_dias['Ruta'] == ruta)].reset_index(drop=True)
                    if not df_unicaruta.empty:
                            lugares_lista = df_unicaruta['Lugar'].tolist()
                            id_ruta = f"{dia} - {ruta}"
                            
                            if tipo_ruteo == "Ruteo Optimizado (IA) v2" and usar_config_global_v2:
                                st.sidebar.markdown(f"**Día/Ruta:** {id_ruta}")
                            else:
                                st.sidebar.markdown(f"**Ruta:** {id_ruta}")
                            
                            hora_salida_rutas_dict[id_ruta] = st.sidebar.time_input("Hora Salida:", datetime.time(9, 0), key=f"hora_{id_ruta}")
                            
                            if tipo_ruteo == "Ruteo Optimizado (IA)":
                                opciones_lugar = ["🤖 IA Decide"] + lugares_lista
                                sel_ini = st.sidebar.selectbox("Punto de Inicio:", opciones_lugar, index=0, key=f"ini_{id_ruta}")
                                sel_anteante = st.sidebar.selectbox("Punto Ante-antepenúltimo:", opciones_lugar, index=0, key=f"anteante_{id_ruta}")
                                sel_ante = st.sidebar.selectbox("Punto Antepenúltimo:", opciones_lugar, index=0, key=f"ante_{id_ruta}")
                                sel_pen = st.sidebar.selectbox("Punto Penúltimo:", opciones_lugar, index=0, key=f"pen_{id_ruta}")
                                sel_fin = st.sidebar.selectbox("Punto Final:", opciones_lugar, index=len(opciones_lugar)-1, key=f"fin_{id_ruta}")
                                
                                opciones_inicio_dict[id_ruta] = sel_ini
                                opciones_anteante_dict[id_ruta] = sel_anteante
                                opciones_antepenultimo_dict[id_ruta] = sel_ante
                                opciones_penultimo_dict[id_ruta] = sel_pen
                                opciones_fin_dict[id_ruta] = sel_fin
                                
                            elif tipo_ruteo == "Ruteo Optimizado (IA) v2":
                                if usar_config_global_v2:
                                    opciones_inicio_dict[id_ruta] = opciones_inicio_global.get(ruta, "🤖 IA Decide")
                                    opciones_deptos_dict[id_ruta] = opciones_deptos_dict_global.get(ruta, {})
                                else:
                                    lugares_lista_limpia = [loc for loc in df_unicaruta['Lugar'].tolist() if str(df_unicaruta[df_unicaruta['Lugar']==loc]['Departamento'].iloc[0]).strip().upper() != 'LABNU']
                                    opciones_lugar = ["🤖 IA Decide"] + lugares_lista_limpia
                                    
                                    sel_ini = st.sidebar.selectbox("Inicio Global:", opciones_lugar, index=0, key=f"ini_v2_{id_ruta}")
                                    opciones_inicio_dict[id_ruta] = sel_ini
                                    
                                    deptos_lista = df_unicaruta['Departamento'].unique().tolist()
                                    opciones_deptos_dict[id_ruta] = {}
                                    
                                    for dept in deptos_lista:
                                        if pd.isna(dept) or str(dept).strip() == '': continue
                                        dept_str = str(dept).strip()
                                        
                                        # IGNORAMOS LABNU EN LA INTERFAZ
                                        if dept_str.upper() == 'LABNU': 
                                            continue
                                            
                                        st.sidebar.markdown(f"🔹 *Depto: {dept_str}*")
                                        l_dept = df_unicaruta[df_unicaruta['Departamento'] == dept]['Lugar'].tolist()
                                        l_dept = [loc for loc in l_dept if str(loc).strip().upper() != 'LABNU']
                                        
                                        opc_dept = ["🤖 IA Decide"] + l_dept
                                        
                                        sel_aa = st.sidebar.selectbox("Ante-antepenúltimo:", opc_dept, index=0, key=f"aa_{id_ruta}_{dept_str}")
                                        sel_a = st.sidebar.selectbox("Antepenúltimo:", opc_dept, index=0, key=f"a_{id_ruta}_{dept_str}")
                                        sel_p = st.sidebar.selectbox("Penúltimo:", opc_dept, index=0, key=f"p_{id_ruta}_{dept_str}")
                                        sel_f = st.sidebar.selectbox("Último:", opc_dept, index=0, key=f"f_{id_ruta}_{dept_str}")
                                        
                                        opciones_deptos_dict[id_ruta][dept_str] = {
                                            'aa': sel_aa, 'a': sel_a, 'p': sel_p, 'f': sel_f
                                        }
                                
                            st.sidebar.markdown("<br>", unsafe_allow_html=True)

        elif "Creación de rutas propias" in tipo_ruteo:
            st.sidebar.markdown("---")
            st.sidebar.header("Configuración de Flota Automática")
            
            if "Patrón Fijo" in tipo_ruteo:
                st.sidebar.info("🗓️ Modo Patrón Maestro (Tu Lógica): Extrae los clientes esporádicos para asegurar una FLOTA MÍNIMA BASE. Luego inyecta esos puntos controlando el reloj para JAMÁS pasar de las 14:30 hs (+10 min de tolerancia estricta).")
            elif "Fijo" in tipo_ruteo:
                st.sidebar.info("🏢 Modo Fijo: Corta el mapa y calcula flota 100% independiente por departamento. NUNCA mezcla zonas en un auto.")
            elif "Flexible" in tipo_ruteo:
                st.sidebar.info("🏘️ Modo Flexible: Agrupa por zona para dar orden, pero SÍ PERMITE cruzar fronteras si eso ahorra crear un vehículo entero.")
            else:
                st.sidebar.info("🚀 Modo Ideal Libre: Ignora fronteras geográficas. Prioriza únicamente el ahorro MÁXIMO de vehículos.")
                
            opciones_lugar_vrp = df_filtrado_dias['Lugar'].unique().tolist() if dias_seleccionados else []
            if not opciones_lugar_vrp:
                st.sidebar.error("No hay lugares válidos en los días seleccionados.")
                st.stop()
                
            punto_final_vrp = st.sidebar.selectbox("📍 Punto final de TODAS las rutas:", opciones_lugar_vrp)
            
            col_salida, col_llegada = st.sidebar.columns(2)
            with col_salida:
                hora_salida_vrp = st.time_input("Hora Salida", datetime.time(8, 0))
            with col_llegada:
                hora_llegada_vrp = st.time_input("Límite Llegada", datetime.time(14, 30))
                
            min_parada_vrp = st.sidebar.number_input("Minutos espera por parada", min_value=0, value=15, step=1)
            
            start_dt = datetime.datetime.combine(datetime.date.today(), hora_salida_vrp)
            end_dt = datetime.datetime.combine(datetime.date.today(), hora_llegada_vrp)
            max_time_sec = int((end_dt - start_dt).total_seconds())
            
            if max_time_sec <= 0:
                st.sidebar.error("❌ El horario de llegada debe ser mayor al de salida.")
                st.stop()

        # --- BOTÓN DE CÁLCULO ---
        if st.sidebar.button("🗺️ Calcular Rutas", type="primary"):
            st.session_state['hora_salida_rutas_dict'] = hora_salida_rutas_dict
            st.session_state['tipo_ruteo'] = tipo_ruteo # <-- BLINDAJE EN MEMORIA PARA MANTENER LA HORA A LAS 09:00
            
            if tipo_ruteo in ["Ruteo según Excel (Orden Original)", "Ruteo Optimizado (IA)", "Ruteo Optimizado (IA) v2"]:
                st.session_state['min_parada_guardado'] = min_parada_global
            else:
                st.session_state['min_parada_guardado'] = min_parada_vrp
                
            with st.spinner("Procesando la red logística y calculando tiempos..."):
                
                lat_centro = df_filtrado_dias.iloc[0]['Coords_Procesadas'][1]
                lon_centro = df_filtrado_dias.iloc[0]['Coords_Procesadas'][0]
                mapa_calculado = folium.Map(location=[lat_centro, lon_centro], zoom_start=11)

                colores = ['blue', 'red', 'green', 'purple', 'orange', 'darkred', 'cadetblue', 'darkblue', 'pink', 'lightgreen']
                datos_para_resumen = []
                color_idx = 0
                headers = {'Authorization': api_key, 'Content-Type': 'application/json'}

                destino_row_global = None
                if "Creación de rutas propias" in tipo_ruteo:
                    try:
                        destino_row_global = df_filtrado_dias[df_filtrado_dias['Lugar'] == punto_final_vrp].iloc[0]
                    except Exception:
                        destino_row_global = df[df['Lugar'] == punto_final_vrp].iloc[0]

                # ==========================================================
                # LÓGICA 1 Y 2: RUTEO CLÁSICO Y OPTIMIZADOS
                # ==========================================================
                if tipo_ruteo in ["Ruteo según Excel (Orden Original)", "Ruteo Optimizado (IA)", "Ruteo Optimizado (IA) v2"]:
                    for dia in dias_seleccionados:
                        df_dia_general = df[df['Día'] == dia]
                        if not df_dia_general.empty:
                            dibujar_geozona_circular(df_dia_general['Coords_Procesadas'].tolist(), f"🌍 DÍA: {dia}", "black", mapa_calculado)

                        for ruta in rutas_seleccionadas:
                            df_ruta = df[(df['Día'] == dia) & (df['Ruta'] == ruta)].copy().reset_index(drop=True)
                            if df_ruta.empty: continue
                            
                            id_unico = f"{dia} - {ruta}"
                            color_actual = colores[color_idx % len(colores)]
                            color_idx += 1
                            
                            lista_coords = df_ruta['Coords_Procesadas'].tolist()
                            nodos_ordenados = []
                            coords_ordenadas = []

                            if tipo_ruteo == "Ruteo según Excel (Orden Original)":
                                nodos_ordenados = list(range(len(df_ruta)))
                                coords_ordenadas = lista_coords
                            else: 
                                num_locs = len(lista_coords)
                                if num_locs < 2:
                                    nodos_ordenados = [0]
                                    coords_ordenadas = lista_coords
                                else:
                                    matriz_dist, matriz_dur, err_matriz = obtener_matriz_masiva(lista_coords, headers)
                                    if err_matriz == "QUOTA_EXCEEDED":
                                        st.error(f"❌ ¡SALDO DIARIO AGOTADO en la ruta {ruta}! Pega tu propia clave en el menú lateral.")
                                        st.stop()
                                        
                                    if not err_matriz:
                                        N = num_locs
                                        extended_dist = [[0] * (N + 2) for _ in range(N + 2)]
                                        
                                        for i in range(N):
                                            for j in range(N):
                                                val = matriz_dist[i][j]
                                                extended_dist[i][j] = int(val) if val is not None else 99999999
                                                
                                        sel_inicio = opciones_inicio_dict.get(id_unico, "🤖 IA Decide")
                                        lugares_actuales = df_ruta['Lugar'].tolist()
                                        
                                        idx_inicio = lugares_actuales.index(sel_inicio) if sel_inicio in lugares_actuales and sel_inicio != "🤖 IA Decide" else -1
                                        
                                        if idx_inicio != -1:
                                            for j in range(N): extended_dist[N][j] = 99999999
                                            extended_dist[N][idx_inicio] = 0
                                        else:
                                            for j in range(N): extended_dist[N][j] = 0
                                            
                                        if tipo_ruteo == "Ruteo Optimizado (IA)":
                                            sel_anteante = opciones_anteante_dict.get(id_unico, "🤖 IA Decide")
                                            sel_ante = opciones_antepenultimo_dict.get(id_unico, "🤖 IA Decide")
                                            sel_pen = opciones_penultimo_dict.get(id_unico, "🤖 IA Decide")
                                            sel_fin = opciones_fin_dict.get(id_unico, "🤖 IA Decide")
                                            
                                            idx_anteante = lugares_actuales.index(sel_anteante) if sel_anteante in lugares_actuales and sel_anteante != "🤖 IA Decide" else -1
                                            idx_ante = lugares_actuales.index(sel_ante) if sel_ante in lugares_actuales and sel_ante != "🤖 IA Decide" else -1
                                            idx_pen = lugares_actuales.index(sel_pen) if sel_pen in lugares_actuales and sel_pen != "🤖 IA Decide" else -1
                                            idx_fin = lugares_actuales.index(sel_fin) if sel_fin in lugares_actuales and sel_fin != "🤖 IA Decide" else -1
                                            
                                            if idx_fin != -1:
                                                for i in range(N): extended_dist[i][N+1] = 99999999
                                                extended_dist[idx_fin][N+1] = 0
                                            else:
                                                for i in range(N): extended_dist[i][N+1] = 0

                                            if idx_pen != -1 and idx_fin != -1:
                                                for j in range(N): 
                                                    if j != idx_fin: extended_dist[idx_pen][j] = 99999999
                                                for i in range(N + 1): 
                                                    if i != idx_pen: extended_dist[i][idx_fin] = 99999999

                                            if idx_ante != -1 and idx_pen != -1:
                                                for j in range(N):
                                                    if j != idx_pen: extended_dist[idx_ante][j] = 99999999
                                                for i in range(N + 1): 
                                                    if i != idx_ante: extended_dist[i][idx_pen] = 99999999
                                                    
                                            if idx_anteante != -1 and idx_ante != -1:
                                                for j in range(N):
                                                    if j != idx_ante: extended_dist[idx_anteante][j] = 99999999
                                                for i in range(N + 1): 
                                                    if i != idx_anteante: extended_dist[i][idx_ante] = 99999999
                                                    
                                        elif tipo_ruteo == "Ruteo Optimizado (IA) v2":
                                            deptos_actuales = df_ruta['Departamento'].tolist()
                                            idx_labnu = -1
                                            for idx_loc, depto_val in enumerate(deptos_actuales):
                                                if str(depto_val).strip().upper() == 'LABNU':
                                                    idx_labnu = idx_loc
                                                    break
                                            
                                            if idx_labnu != -1:
                                                for i in range(N): extended_dist[i][N+1] = 99999999
                                                extended_dist[idx_labnu][N+1] = 0
                                            else:
                                                for i in range(N): extended_dist[i][N+1] = 0

                                        manager = pywrapcp.RoutingIndexManager(N + 2, 1, [N], [N+1])
                                        routing = pywrapcp.RoutingModel(manager)
                                        
                                        def distance_callback(from_index, to_index):
                                            from_node = manager.IndexToNode(from_index)
                                            to_node = manager.IndexToNode(to_index)
                                            return int(extended_dist[from_node][to_node])
                                            
                                        transit_callback_index = routing.RegisterTransitCallback(distance_callback)
                                        routing.SetArcCostEvaluatorOfAllVehicles(transit_callback_index)
                                        
                                        if tipo_ruteo == "Ruteo Optimizado (IA) v2":
                                            def sequence_callback(from_index):
                                                return 1
                                            seq_call_idx = routing.RegisterUnaryTransitCallback(sequence_callback)
                                            routing.AddDimension(seq_call_idx, 0, 9999, True, "Sequence")
                                            seq_dim = routing.GetDimensionOrDie("Sequence")
                                            solver = routing.solver()
                                            
                                            deptos_actuales = df_ruta['Departamento'].tolist()
                                            dept_config = opciones_deptos_dict.get(id_unico, {})
                                            
                                            for dept_str, config in dept_config.items():
                                                sel_aa = config['aa']
                                                sel_a = config['a']
                                                sel_p = config['p']
                                                sel_f = config['f']
                                                
                                                idx_aa = lugares_actuales.index(sel_aa) if sel_aa in lugares_actuales and sel_aa != "🤖 IA Decide" else -1
                                                idx_a = lugares_actuales.index(sel_a) if sel_a in lugares_actuales and sel_a != "🤖 IA Decide" else -1
                                                idx_p = lugares_actuales.index(sel_p) if sel_p in lugares_actuales and sel_p != "🤖 IA Decide" else -1
                                                idx_f = lugares_actuales.index(sel_f) if sel_f in lugares_actuales and sel_f != "🤖 IA Decide" else -1
                                                
                                                target_last_nodes = []
                                                for x in [idx_aa, idx_a, idx_p, idx_f]:
                                                    if x != -1 and x not in target_last_nodes and x != idx_inicio and x != idx_labnu:
                                                        target_last_nodes.append(x)
                                                        
                                                special_indices = set(target_last_nodes)
                                                reg_indices = [i for i in range(N) if str(deptos_actuales[i]).strip() == dept_str and i not in special_indices and i != idx_inicio and i != idx_labnu]
                                                
                                                if len(target_last_nodes) > 0:
                                                    first_special = target_last_nodes[0]
                                                    for r in reg_indices:
                                                        solver.Add(seq_dim.CumulVar(manager.NodeToIndex(r)) < seq_dim.CumulVar(manager.NodeToIndex(first_special)))
                                                        
                                                for i in range(len(target_last_nodes) - 1):
                                                    node_before = target_last_nodes[i]
                                                    node_after = target_last_nodes[i+1]
                                                    solver.Add(seq_dim.CumulVar(manager.NodeToIndex(node_before)) < seq_dim.CumulVar(manager.NodeToIndex(node_after)))
                                        
                                        search_parameters = pywrapcp.DefaultRoutingSearchParameters()
                                        # VOLVIMOS A SAVINGS: El motor más estable y rápido. No se frustra con las secuencias lógicas.
                                        search_parameters.first_solution_strategy = routing_enums_pb2.FirstSolutionStrategy.SAVINGS
                                        search_parameters.local_search_metaheuristic = routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH
                                        search_parameters.time_limit.seconds = 5 
                                        
                                        solution = routing.SolveWithParameters(search_parameters)
                                        
                                        if solution:
                                            idx = routing.Start(0)
                                            while not routing.IsEnd(idx):
                                                node = manager.IndexToNode(idx)
                                                if node < N:
                                                    nodos_ordenados.append(node)
                                                idx = solution.Value(routing.NextVar(idx))
                                            coords_ordenadas = [lista_coords[i] for i in nodos_ordenados]
                                        else:
                                            st.error(f"No se encontró solución matemática para {ruta}. Revisa no haber creado un bucle imposible.")
                                            continue
                                    else:
                                        st.error(f"Error Matriz en {ruta}: {err_matriz}")
                                        continue

                            if len(coords_ordenadas) > 1:
                                geojson, err_dirs = obtener_trazado_masivo(coords_ordenadas, headers)
                                if err_dirs == "QUOTA_EXCEEDED":
                                    st.error(f"❌ ¡SALDO DIARIO AGOTADO al dibujar {ruta}!")
                                    st.stop()
                                    
                                if not err_dirs:
                                    props = geojson['features'][0]['properties']['summary']
                                    segments = geojson['features'][0]['properties'].get('segments', [])
                                    
                                    paradas_info = []
                                    for nodo_idx in nodos_ordenados:
                                        fila = df_ruta.iloc[nodo_idx]
                                        paradas_info.append({
                                            "Día": fila.get('Día',''), "Ruta": fila.get('Ruta',''),
                                            "Departamento": fila.get('Departamento',''), "Lugar": fila.get('Lugar',''),
                                            "Coordenadas": fila.get('Coordenadas','')
                                        })

                                    datos_para_resumen.append({
                                        "id_unico": id_unico, "dia": dia, "ruta": ruta,
                                        "puntos": len(df_ruta), "dist_km": round(props['distance']/1000, 2),
                                        "drive_mins": round(props['duration']/60, 0), "color": color_actual,
                                        "paradas": paradas_info, "segmentos": segments,
                                        "geojson": geojson,
                                        "coords_ordenadas": coords_ordenadas
                                    })
                                else:
                                    st.error(f"Error trazando calles de {ruta}: {err_dirs}")

                # ==========================================================
                # LÓGICA 3 Y 4: CREACIÓN DE RUTAS PROPIAS (LIBRE Y FLEXIBLE DIARIO)
                # ==========================================================
                elif tipo_ruteo in ["Creación de rutas propias (Ideal Libre)", "Creación de rutas propias (Departamental Flexible)"]:
                    for dia in dias_seleccionados:
                        df_dia = df[df['Día'] == dia].copy().reset_index(drop=True)
                        if punto_final_vrp not in df_dia['Lugar'].values:
                            df_dia = pd.concat([df_dia, destino_row_global.to_frame().T], ignore_index=True)
                        
                        end_idx = df_dia[df_dia['Lugar'] == punto_final_vrp].index[0]
                        lista_coords = df_dia['Coords_Procesadas'].tolist()
                        num_locs = len(lista_coords)
                        
                        if num_locs < 2: continue

                        matriz_dist, matriz_dur, err_matriz = obtener_matriz_masiva(lista_coords, headers)
                        if err_matriz == "QUOTA_EXCEEDED":
                            st.error(f"❌ ¡SALDO DIARIO AGOTADO en el Día {dia}!")
                            st.stop()
                            
                        if not err_matriz:
                            dummy_idx = num_locs
                            for i in range(num_locs):
                                matriz_dist[i].append(0); matriz_dur[i].append(0)
                            matriz_dist.append([0] * (num_locs + 1)); matriz_dur.append([0] * (num_locs + 1))
                            
                            for i in range(num_locs):
                                if i != dummy_idx and i != end_idx:
                                    val_dur = matriz_dur[i][end_idx]
                                    if val_dur is None:
                                        st.error(f"❌ Error de Mapa: El punto '{df_dia.iloc[i]['Lugar']}' no tiene conexión por calle.")
                                        st.stop()
                                    tiempo_minimo_viaje = int(min_parada_vrp * 60) + int(val_dur)
                                    if tiempo_minimo_viaje > max_time_sec:
                                        st.error(f"❌ Error Físico Real: Ir desde '{df_dia.iloc[i]['Lugar']}' hasta el destino final toma por sí solo {tiempo_minimo_viaje//60} min.")
                                        st.stop()

                            num_vehicles = num_locs 
                            manager = pywrapcp.RoutingIndexManager(num_locs + 1, num_vehicles, [dummy_idx] * num_vehicles, [int(end_idx)] * num_vehicles)
                            routing = pywrapcp.RoutingModel(manager)
                            
                            def distance_callback(from_index, to_index):
                                from_node = manager.IndexToNode(from_index)
                                to_node = manager.IndexToNode(to_index)
                                val = matriz_dist[from_node][to_node]
                                dist = int(val) if val is not None else 99999999 
                                
                                if from_node == dummy_idx and to_node != end_idx:
                                    return dist + 100000000 
                                    
                                if "Flexible" in tipo_ruteo:
                                    if from_node < num_locs and to_node < num_locs and from_node != end_idx and to_node != end_idx and from_node != dummy_idx:
                                        dept_f = str(df_dia.iloc[from_node].get('Departamento', '')).strip().lower()
                                        dept_t = str(df_dia.iloc[to_node].get('Departamento', '')).strip().lower()
                                        if dept_f and dept_t and dept_f != dept_t:
                                            dist += 500000 
                                return int(dist)
                                
                            transit_callback_index = routing.RegisterTransitCallback(distance_callback)
                            routing.SetArcCostEvaluatorOfAllVehicles(transit_callback_index)
                            
                            routing.SetFixedCostOfAllVehicles(100000000)
                            
                            def time_callback(from_index, to_index):
                                from_node = manager.IndexToNode(from_index)
                                to_node = manager.IndexToNode(to_index)
                                val_dur = matriz_dur[from_node][to_node]
                                drive_time = int(val_dur) if val_dur is not None else 99999999 
                                wait_time = int(min_parada_vrp * 60) if to_node != dummy_idx and to_node != end_idx else 0
                                return int(drive_time + wait_time)
                                
                            time_callback_index = routing.RegisterTransitCallback(time_callback)
                            routing.AddDimension(time_callback_index, 0, max_time_sec, True, "Time")

                            search_parameters = pywrapcp.DefaultRoutingSearchParameters()
                            search_parameters.first_solution_strategy = routing_enums_pb2.FirstSolutionStrategy.SAVINGS
                            search_parameters.time_limit.seconds = 20 
                            
                            solution = routing.SolveWithParameters(search_parameters)
                            
                            if solution:
                                vehiculo_real_count = 1
                                for vehicle_id in range(num_vehicles):
                                    index = routing.Start(vehicle_id)
                                    first_visit = solution.Value(routing.NextVar(index))
                                    
                                    if manager.IndexToNode(first_visit) == end_idx:
                                        continue 
                                    
                                    nodos_ordenados = []
                                    while not routing.IsEnd(index):
                                        node = manager.IndexToNode(index)
                                        if node != dummy_idx:
                                            nodos_ordenados.append(node)
                                        index = solution.Value(routing.NextVar(index))
                                    nodos_ordenados.append(end_idx)
                                    
                                    coords_ordenadas = [lista_coords[i] for i in nodos_ordenados]
                                    ruta_nombre = f"Auto {vehiculo_real_count}"
                                    id_unico = f"{dia} - {ruta_nombre}"
                                    vehiculo_real_count += 1
                                    color_actual = colores[color_idx % len(colores)]
                                    color_idx += 1
                                    
                                    geojson, err_dirs = obtener_trazado_masivo(coords_ordenadas, headers)
                                    if err_dirs == "QUOTA_EXCEEDED":
                                        st.error(f"❌ ¡SALDO DIARIO AGOTADO al dibujar {ruta_nombre}!")
                                        st.stop()
                                        
                                    if not err_dirs:
                                        props = geojson['features'][0]['properties']['summary']
                                        segments = geojson['features'][0]['properties'].get('segments', [])
                                        paradas_info = []
                                        for nodo_idx in nodos_ordenados:
                                            fila = df_dia.iloc[nodo_idx]
                                            paradas_info.append({
                                                "Día": fila.get('Día',''), "Ruta": ruta_nombre, 
                                                "Departamento": fila.get('Departamento',''), "Lugar": fila.get('Lugar',''),
                                                "Coordenadas": fila.get('Coordenadas','')
                                            })
                                        datos_para_resumen.append({
                                            "id_unico": id_unico, "dia": dia, "ruta": ruta_nombre,
                                            "puntos": len(nodos_ordenados), "dist_km": round(props['distance']/1000, 2),
                                            "drive_mins": round(props['duration']/60, 0), "color": color_actual,
                                            "paradas": paradas_info, "segmentos": segments,
                                            "geojson": geojson,
                                            "coords_ordenadas": coords_ordenadas
                                        })
                                    else:
                                        st.error(f"Error en trazado: {err_dirs}")
                            else:
                                st.error(f"❌ Imposible matemático en el {dia}.")
                        else:
                            st.error(f"Error Matriz {dia}: {err_matriz}")

                # ==========================================================
                # LÓGICA 5: CREACIÓN DE RUTAS PROPIAS (DEPARTAMENTAL FIJO - NORMAL)
                # ==========================================================
                elif tipo_ruteo == "Creación de rutas propias (Departamental Fijo)":
                    for dia in dias_seleccionados:
                        df_dia_completo = df[df['Día'] == dia].copy().reset_index(drop=True)
                        dept_series = df_dia_completo[df_dia_completo['Lugar'] != punto_final_vrp]['Departamento']
                        departamentos = [d for d in dept_series.unique() if pd.notna(d) and str(d).strip() != '']
                        vehiculo_real_count = 1
                        
                        for dept in departamentos:
                            df_dept = df_dia_completo[(df_dia_completo['Departamento'] == dept) & (df_dia_completo['Lugar'] != punto_final_vrp)].copy().reset_index(drop=True)
                            if df_dept.empty: continue
                            
                            df_dept = pd.concat([df_dept, destino_row_global.to_frame().T], ignore_index=True)
                            end_idx = len(df_dept) - 1
                            lista_coords = df_dept['Coords_Procesadas'].tolist()
                            num_locs = len(lista_coords)
                            
                            if num_locs < 2: continue

                            matriz_dist, matriz_dur, err_matriz = obtener_matriz_masiva(lista_coords, headers)
                            if err_matriz == "QUOTA_EXCEEDED":
                                st.error(f"❌ ¡SALDO DIARIO AGOTADO en el Día {dia}, Depto {dept}!")
                                st.stop()
                                
                            if not err_matriz:
                                dummy_idx = num_locs
                                for i in range(num_locs):
                                    matriz_dist[i].append(0); matriz_dur[i].append(0)
                                matriz_dist.append([0] * (num_locs + 1)); matriz_dur.append([0] * (num_locs + 1))
                                
                                for i in range(num_locs):
                                    if i != dummy_idx and i != end_idx:
                                        val_dur = matriz_dur[i][end_idx]
                                        if val_dur is None:
                                            st.error(f"❌ Error de Mapa: El punto '{df_dept.iloc[i]['Lugar']}' no tiene conexión por calle.")
                                            st.stop()
                                        tiempo_minimo_viaje = int(min_parada_vrp * 60) + int(val_dur)
                                        if tiempo_minimo_viaje > max_time_sec:
                                            st.error(f"❌ Error Físico Real: Ir desde '{df_dept.iloc[i]['Lugar']}' ({dept}) hasta el destino toma {tiempo_minimo_viaje//60} min reales.")
                                            st.stop()

                                num_vehicles = num_locs 
                                manager = pywrapcp.RoutingIndexManager(num_locs + 1, num_vehicles, [dummy_idx] * num_vehicles, [int(end_idx)] * num_vehicles)
                                routing = pywrapcp.RoutingModel(manager)
                                
                                def distance_callback(from_index, to_index):
                                    from_node = manager.IndexToNode(from_index)
                                    to_node = manager.IndexToNode(to_index)
                                    val = matriz_dist[from_node][to_node]
                                    dist = int(val) if val is not None else 99999999 
                                    if from_node == dummy_idx and to_node != end_idx: return dist + 100000000 
                                    return int(dist)
                                    
                                transit_callback_index = routing.RegisterTransitCallback(distance_callback)
                                routing.SetArcCostEvaluatorOfAllVehicles(transit_callback_index)
                                
                                routing.SetFixedCostOfAllVehicles(100000000)
                                
                                def time_callback(from_index, to_index):
                                    from_node = manager.IndexToNode(from_index)
                                    to_node = manager.IndexToNode(to_index)
                                    val_dur = matriz_dur[from_node][to_node]
                                    drive_time = int(val_dur) if val_dur is not None else 99999999 
                                    wait_time = int(min_parada_vrp * 60) if to_node != dummy_idx and to_node != end_idx else 0
                                    return int(drive_time + wait_time)
                                    
                                time_callback_index = routing.RegisterTransitCallback(time_callback)
                                routing.AddDimension(time_callback_index, 0, max_time_sec, True, "Time")

                                search_parameters = pywrapcp.DefaultRoutingSearchParameters()
                                search_parameters.first_solution_strategy = routing_enums_pb2.FirstSolutionStrategy.SAVINGS
                                search_parameters.local_search_metaheuristic = routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH
                                search_parameters.time_limit.seconds = 10 
                                
                                solution = routing.SolveWithParameters(search_parameters)
                                
                                if solution:
                                    for vehicle_id in range(num_vehicles):
                                        index = routing.Start(vehicle_id)
                                        first_visit = solution.Value(routing.NextVar(index))
                                        if manager.IndexToNode(first_visit) == end_idx: continue 
                                        nodos_ordenados = []
                                        while not routing.IsEnd(index):
                                            node = manager.IndexToNode(index)
                                            if node != dummy_idx: nodos_ordenados.append(node)
                                            index = solution.Value(routing.NextVar(index))
                                        nodos_ordenados.append(end_idx)
                                        
                                        coords_ordenadas = [lista_coords[i] for i in nodos_ordenados]
                                        nombre_dept_limpio = str(dept).strip()
                                        ruta_nombre = f"Auto {vehiculo_real_count} ({nombre_dept_limpio})"
                                        id_unico = f"{dia} - {ruta_nombre}"
                                        vehiculo_real_count += 1
                                        color_actual = colores[color_idx % len(colores)]
                                        color_idx += 1
                                        
                                        geojson, err_dirs = obtener_trazado_masivo(coords_ordenadas, headers)
                                        if err_dirs == "QUOTA_EXCEEDED":
                                            st.error(f"❌ ¡SALDO DIARIO AGOTADO al dibujar {ruta_nombre}!")
                                            st.stop()
                                        if not err_dirs:
                                            props = geojson['features'][0]['properties']['summary']
                                            segments = geojson['features'][0]['properties'].get('segments', [])
                                            paradas_info = []
                                            for nodo_idx in nodos_ordenados:
                                                fila = df_dept.iloc[nodo_idx]
                                                paradas_info.append({
                                                    "Día": fila.get('Día',''), "Ruta": ruta_nombre, 
                                                    "Departamento": fila.get('Departamento',''), "Lugar": fila.get('Lugar',''),
                                                    "Coordenadas": fila.get('Coordenadas','')
                                                })
                                            datos_para_resumen.append({
                                                "id_unico": id_unico, "dia": dia, "ruta": ruta_nombre,
                                                "puntos": len(nodos_ordenados), "dist_km": round(props['distance']/1000, 2),
                                                "drive_mins": round(props['duration']/60, 0), "color": color_actual,
                                                "paradas": paradas_info, "segmentos": segments,
                                                "geojson": geojson,
                                                "coords_ordenadas": coords_ordenadas
                                            })
                                        else:
                                            st.error(f"Error en trazado: {err_dirs}")
                                else:
                                    st.error(f"❌ Imposible matemático en el {dia} para {dept}.")
                            else:
                                st.error(f"Error Matriz {dia} - {dept}: {err_matriz}")

                # ==========================================================
                # LÓGICA 6, 7 Y 8: CREACIÓN DE RUTAS PROPIAS (PATRÓN MAESTRO - CON LÍMITE ESTRICTO DE +10 MIN)
                # ==========================================================
                elif "Patrón Fijo" in tipo_ruteo:
                    st.info("🧠 Generando Patrón Maestro: Extrayendo los clientes esporádicos para garantizar la FLOTA MÍNIMA. Los sobrantes se inyectarán con límite matemático ESTRICTO de +10 minutos.")
                    
                    df_total_puntos = df_filtrado_dias[df_filtrado_dias['Lugar'] != punto_final_vrp].drop_duplicates(subset=['Lugar']).copy().reset_index(drop=True)
                    
                    # MATRIZ ÚNICA MUNDIAL
                    df_master_global = pd.concat([df_total_puntos, destino_row_global.to_frame().T], ignore_index=True)
                    lista_coords_global = df_master_global['Coords_Procesadas'].tolist()
                    lugares_globales = df_master_global['Lugar'].tolist()
                    end_idx_global = len(lugares_globales) - 1

                    matriz_dist_global, matriz_dur_global, err_matriz_glob = obtener_matriz_masiva(lista_coords_global, headers)
                    if err_matriz_glob: st.error(err_matriz_glob); st.stop()
                    
                    # LIMPIEZA MATEMÁTICA A ENTEROS
                    for i in range(len(matriz_dist_global)):
                        for j in range(len(matriz_dist_global[0])):
                            if matriz_dist_global[i][j] is None: matriz_dist_global[i][j] = 99999999
                            else: matriz_dist_global[i][j] = int(matriz_dist_global[i][j])
                            if matriz_dur_global[i][j] is None: matriz_dur_global[i][j] = 99999999
                            else: matriz_dur_global[i][j] = int(matriz_dur_global[i][j])

                    def calcular_tiempo_ruta_en_dia_especifico(ruta_locs, day):
                        locs_day = set(df_filtrado_dias[df_filtrado_dias['Día'] == day]['Lugar'])
                        r_day = [l for l in ruta_locs if l in locs_day]
                        if not r_day: return 0, 0
                        
                        t = 0
                        d = 0
                        for i in range(len(r_day) - 1):
                            idx_A = lugares_globales.index(r_day[i])
                            idx_B = lugares_globales.index(r_day[i+1])
                            t += int(min_parada_vrp * 60) + matriz_dur_global[idx_A][idx_B]
                            d += matriz_dist_global[idx_A][idx_B]
                            
                        idx_last = lugares_globales.index(r_day[-1])
                        t += int(min_parada_vrp * 60) + matriz_dur_global[idx_last][end_idx_global]
                        d += matriz_dist_global[idx_last][end_idx_global]
                        return t, d

                    rutas_maestras_base = []
                    vehiculo_real_count = 1

                    subsets = []
                    if "Departamental Fijo" in tipo_ruteo:
                        departamentos = [d for d in df_total_puntos['Departamento'].unique() if pd.notna(d) and str(d).strip() != '']
                        for dept in departamentos:
                            subsets.append((dept, df_total_puntos[df_total_puntos['Departamento'] == dept]['Lugar'].tolist()))
                    else:
                        subsets.append(("General", df_total_puntos['Lugar'].tolist()))

                    for subset_name, subset_lugares in subsets:
                        # 1. EL FILTRO DE FRECUENCIA DEL USUARIO
                        df_subset_filtrado = df_filtrado_dias[df_filtrado_dias['Lugar'].isin(subset_lugares)]
                        dia_pico = df_subset_filtrado.groupby('Día').size().idxmax()
                        lugares_pico_base = df_subset_filtrado[df_subset_filtrado['Día'] == dia_pico]['Lugar'].unique().tolist()
                        
                        pico_indices = [lugares_globales.index(l) for l in lugares_pico_base]
                        num_sub = len(pico_indices)
                        dummy_idx = num_sub
                        end_idx_sub = num_sub + 1
                        
                        rutas_core_locs = []
                        
                        if num_sub < 2:
                            if num_sub == 1: rutas_core_locs.append([lugares_pico_base[0]])
                        else:
                            sub_dist = [[0]*(num_sub+2) for _ in range(num_sub+2)]
                            sub_dur = [[0]*(num_sub+2) for _ in range(num_sub+2)]
                            for i in range(num_sub):
                                for j in range(num_sub):
                                    sub_dist[i][j] = int(matriz_dist_global[pico_indices[i]][pico_indices[j]])
                                    sub_dur[i][j] = int(matriz_dur_global[pico_indices[i]][pico_indices[j]])
                                sub_dist[i][end_idx_sub] = int(matriz_dist_global[pico_indices[i]][end_idx_global])
                                sub_dur[i][end_idx_sub] = int(matriz_dur_global[pico_indices[i]][end_idx_global])
                                
                            manager = pywrapcp.RoutingIndexManager(num_sub + 2, num_sub, [dummy_idx]*num_sub, [end_idx_sub]*num_sub)
                            routing = pywrapcp.RoutingModel(manager)
                            
                            def d_call(f, t):
                                fn, tn = manager.IndexToNode(f), manager.IndexToNode(t)
                                dist = int(sub_dist[fn][tn])
                                if fn == dummy_idx and tn != end_idx_sub: return dist + 100000000 
                                
                                if "Flexible" in tipo_ruteo and fn < num_sub and tn < num_sub:
                                    g_fn, g_tn = pico_indices[fn], pico_indices[tn]
                                    d_f = str(df_master_global.iloc[g_fn].get('Departamento','')).strip().lower()
                                    d_t = str(df_master_global.iloc[g_tn].get('Departamento','')).strip().lower()
                                    if d_f and d_t and d_f != d_t: dist += 500000
                                return dist
                                
                            def t_call(f, t):
                                fn, tn = manager.IndexToNode(f), manager.IndexToNode(t)
                                wt = int(min_parada_vrp*60) if tn != dummy_idx and tn != end_idx_sub else 0
                                return int(sub_dur[fn][tn]) + wt
                                
                            transit_cb = routing.RegisterTransitCallback(d_call)
                            routing.SetArcCostEvaluatorOfAllVehicles(transit_callback_index)
                            time_cb = routing.RegisterTransitCallback(t_call)
                            
                            routing.AddDimension(time_cb, 0, max_time_sec, True, "Time")
                            routing.SetFixedCostOfAllVehicles(100000000) 
                            
                            search_params = pywrapcp.DefaultRoutingSearchParameters()
                            search_params.first_solution_strategy = routing_enums_pb2.FirstSolutionStrategy.SAVINGS
                            search_params.time_limit.seconds = 10
                            
                            sol = routing.SolveWithParameters(search_params)
                            if sol:
                                for vid in range(num_sub):
                                    idx = routing.Start(vid)
                                    fv = sol.Value(routing.NextVar(idx))
                                    if manager.IndexToNode(fv) == end_idx_sub: continue
                                    
                                    r = []
                                    while not routing.IsEnd(idx):
                                        n = manager.IndexToNode(idx)
                                        if n != dummy_idx: r.append(lugares_pico_base[n])
                                        idx = sol.Value(routing.NextVar(idx))
                                    rutas_core_locs.append(r)
                                    
                        # 2. INYECCIÓN CON RELOJ ESTRICTO (+10 MIN EXACTOS): 
                        lugares_faltantes = [l for l in subset_lugares if l not in lugares_pico_base]
                        
                        for f_loc in lugares_faltantes:
                            dias_activos = df_subset_filtrado[df_subset_filtrado['Lugar'] == f_loc]['Día'].unique()
                            f_idx = lugares_globales.index(f_loc)
                            
                            best_r = -1
                            best_pos = -1
                            min_extra_dist = float('inf')
                            
                            for r_idx, ruta in enumerate(rutas_core_locs):
                                for pos in range(len(ruta) + 1):
                                    ruta_simulada = list(ruta)
                                    ruta_simulada.insert(pos, f_loc)
                                    
                                    es_valida = True
                                    dist_extra_total = 0
                                    
                                    for day in dias_activos:
                                        t_sim, d_sim = calcular_tiempo_ruta_en_dia_especifico(ruta_simulada, day)
                                        
                                        # LA REGLA DE ORO DE LOS +10 MINUTOS EXACTOS (600 SEGUNDOS)
                                        if t_sim > (max_time_sec + 600):
                                            es_valida = False
                                            break
                                        
                                        _, d_orig = calcular_tiempo_ruta_en_dia_especifico(ruta, day)
                                        dist_extra_total += (d_sim - d_orig)
                                        
                                    if es_valida and dist_extra_total < min_extra_dist:
                                        min_extra_dist = dist_extra_total
                                        best_r = r_idx
                                        best_pos = pos
                                        
                            if best_r != -1:
                                rutas_core_locs[best_r].insert(best_pos, f_loc)
                            else:
                                # Si es físicamente imposible meterlo en los autos existentes con +10 min, abre auto nuevo.
                                rutas_core_locs.append([f_loc])
                                
                        for r in rutas_core_locs:
                            name_suffix = f" ({str(subset_name).strip()})" if subset_name != "General" else ""
                            rutas_maestras_base.append({
                                "nombre": f"Auto {vehiculo_real_count}{name_suffix}",
                                "lugares": r,
                                "color_idx": vehiculo_real_count-1
                            })
                            vehiculo_real_count += 1

                    # --- 4. APLICAR EL PATRÓN A CADA DÍA ---
                    st.info("🗓️ Imprimiendo el Patrón Maestro final. Como la IA simuló el reloj estricto en la inyección (Max +10 min), mantenemos la flota mínima sin desfasar los horarios.")
                    for dia in dias_seleccionados:
                        df_dia = df[df['Día'] == dia].copy().reset_index(drop=True)
                        if punto_final_vrp not in df_dia['Lugar'].values:
                            df_dia = pd.concat([df_dia, destino_row_global.to_frame().T], ignore_index=True)

                        lugares_del_dia = set(df_dia['Lugar'].tolist())
                        
                        lista_coords_dia = df_dia['Coords_Procesadas'].tolist()
                        if len(lista_coords_dia) > 1:
                            dibujar_geozona_circular(lista_coords_dia, f"🌍 DÍA: {dia}", "black", mapa_calculado)

                        for ruta_maestra in rutas_maestras_base:
                            lugares_hoy = [l for l in ruta_maestra['lugares'] if l in lugares_del_dia]

                            if len(lugares_hoy) < 2: continue 

                            if lugares_hoy[-1] != punto_final_vrp:
                                if punto_final_vrp in lugares_hoy:
                                    lugares_hoy.remove(punto_final_vrp)
                                lugares_hoy.append(punto_final_vrp)

                            filas_hoy = []
                            coords_ordenadas = []
                            for l in lugares_hoy:
                                fila = df_dia[df_dia['Lugar'] == l].iloc[0]
                                filas_hoy.append(fila)
                                coords_ordenadas.append(fila['Coords_Procesadas'])

                            df_ruta_hoy = pd.DataFrame(filas_hoy)
                            ruta_nombre = ruta_maestra["nombre"]
                            id_unico = f"{dia} - {ruta_nombre}"
                            color_actual = colores[ruta_maestra["color_idx"] % len(colores)]

                            geojson, err_dirs = obtener_trazado_masivo(coords_ordenadas, headers)
                            if err_dirs:
                                st.error(err_dirs)
                                continue

                            props = geojson['features'][0]['properties']['summary']
                            segments = geojson['features'][0]['properties'].get('segments', [])

                            paradas_info = []
                            for _, row_hoy in df_ruta_hoy.iterrows():
                                paradas_info.append({
                                    "Día": row_hoy.get('Día',''), "Ruta": ruta_nombre,
                                    "Departamento": row_hoy.get('Departamento',''), "Lugar": row_hoy.get('Lugar',''),
                                    "Coordenadas": row_hoy.get('Coordenadas','')
                                })

                            datos_para_resumen.append({
                                "id_unico": id_unico, "dia": dia, "ruta": ruta_nombre,
                                "puntos": len(df_ruta_hoy), "dist_km": round(props['distance']/1000, 2),
                                "drive_mins": round(props['duration']/60, 0), "color": color_actual,
                                "paradas": paradas_info, "segmentos": segments,
                                "geojson": geojson,
                                "coords_ordenadas": coords_ordenadas
                            })
                            
                            fg_trazado = folium.FeatureGroup(name=f"🛣️ Trazado: {ruta_nombre} ({dia})")
                            folium.GeoJson(geojson, style_function=lambda x, c=color_actual: {'color':c, 'weight':4, 'opacity':0.8}).add_to(fg_trazado)

                            for i, (_, row_hoy) in enumerate(df_ruta_hoy.iterrows()):
                                lat, lon = row_hoy['Coords_Procesadas'][1], row_hoy['Coords_Procesadas'][0]
                                popup_txt = f"<b>{i+1}. {row_hoy.get('Lugar','')}</b><br>{row_hoy.get('Departamento','')}"
                                icon_html = f"<div style='background:{color_actual};color:white;border-radius:50%;width:20px;text-align:center;border:1px solid white;font-weight:bold;font-size:10pt'>{i+1}</div>"
                                folium.Marker([lat, lon], popup=popup_txt, icon=folium.DivIcon(html=icon_html)).add_to(fg_trazado)

                            fg_trazado.add_to(mapa_calculado)

                folium.LayerControl(collapsed=True).add_to(mapa_calculado)
                mapa_calculado.get_root().html.add_child(folium.Element(js_toggle_capas))
                st.session_state['mapa_guardado'] = mapa_calculado
                st.session_state['datos_resumen'] = datos_para_resumen
                st.session_state['calculo_terminado'] = True

# --- VISTA DE RESULTADOS CON PESTAÑAS (COMÚN PARA TODOS LOS FLUJOS) ---
if st.session_state.get('calculo_terminado', False):
    
    tab_mapa, tab_cronogramas, tab_resumen = st.tabs([
        "🗺️ Mapa Interactivo", 
        "⏱️ Cronogramas Detallados", 
        "📊 Resumen General"
    ])
    
    with tab_mapa:
        st.markdown("### Mapa de Recorridos")
        
        opcion_capas = st.radio(
            "Visibilidad de las rutas:", 
            ["✅ Mostrar Todas las Rutas", "❌ Ocultar Todas las Rutas"], 
            horizontal=True
        )
        mostrar_capas = (opcion_capas == "✅ Mostrar Todas las Rutas")
        
        if st.session_state['datos_resumen']:
            coords_centro = st.session_state['datos_resumen'][0]['coords_ordenadas'][0]
            lat_centro = coords_centro[1]
            lon_centro = coords_centro[0]
            mapa_dinamico = folium.Map(location=[lat_centro, lon_centro], zoom_start=11)
            
            coords_por_dia = {}
            for d in st.session_state['datos_resumen']:
                dia = d['dia']
                if dia not in coords_por_dia: coords_por_dia[dia] = []
                coords_por_dia[dia].extend(d['coords_ordenadas'])
                
            for dia, coords in coords_por_dia.items():
                dibujar_geozona_circular(coords, f"🌍 DÍA: {dia}", "black", mapa_dinamico, mostrar_por_defecto=mostrar_capas)
                
            for d in st.session_state['datos_resumen']:
                dibujar_geozona_circular(d['coords_ordenadas'], f"🗺️ Zona: {d['ruta']} ({d['dia']})", d['color'], mapa_dinamico, mostrar_por_defecto=mostrar_capas)
                
                fg_trazado = folium.FeatureGroup(name=f"🛣️ Trazado: {d['ruta']} ({d['dia']})", show=mostrar_capas)
                folium.GeoJson(d['geojson'], style_function=lambda x, c=d['color']: {'color':c, 'weight':4, 'opacity':0.8}).add_to(fg_trazado)
                
                for i, p in enumerate(d['paradas']):
                    c_lon = d['coords_ordenadas'][i][0]
                    c_lat = d['coords_ordenadas'][i][1]
                    orden_val = p.get('Orden', i+1)
                    popup_txt = f"<b>{orden_val}. {p.get('Lugar','')}</b><br>{p.get('Departamento','')}"
                    icon_html = f"<div style='background:{d['color']};color:white;border-radius:50%;width:20px;text-align:center;border:1px solid white;font-weight:bold;font-size:10pt'>{orden_val}</div>"
                    folium.Marker([c_lat, c_lon], popup=popup_txt, icon=folium.DivIcon(html=icon_html)).add_to(fg_trazado)
                    
                fg_trazado.add_to(mapa_dinamico)
                
            folium.LayerControl(collapsed=True).add_to(mapa_dinamico)
            st_folium(mapa_dinamico, width=1000, height=650, returned_objects=[], key=f"map_dinamico_{mostrar_capas}")
    
    data_global_detallada = []
    data_resumen_general = []
        
    with tab_cronogramas:
        for d in st.session_state['datos_resumen']:
            with st.container():
                st.markdown(f"**📍 {d['dia']} | {d['ruta']}** ({d['puntos']} pts | {d['dist_km']} km)")
                
                c1, c2, c3 = st.columns([1.2, 1, 1])
                
                default_h = datetime.time(9,0)
                default_wait = st.session_state.get('min_parada_guardado', 15)
                
                dict_horas = st.session_state.get('hora_salida_rutas_dict', {})
                if d['id_unico'] in dict_horas:
                    default_h = dict_horas[d['id_unico']]
                
                h_inicio = c1.time_input("Salida", default_h, key=f"h_{d['id_unico']}")
                espera = c2.number_input("Espera por parada (min)", 0, 300, default_wait, key=f"w_{d['id_unico']}")
                
                total_espera_calc = espera * (d['puntos'] - 1) if d['puntos'] > 0 else 0
                total_min = d['drive_mins'] + total_espera_calc
                c3.metric("Total Estimado", f"{total_min:.0f} min")
                
                t_actual = datetime.datetime.combine(datetime.date.today(), h_inicio)
                inicio_dt = t_actual
                dist_acum = 0.0
                mins_acum = 0.0
                rows_excel = []
                hora_llegada_final = "-"
                llegada_final_dt = None
                
                for i, p in enumerate(d['paradas']):
                    dist_tramo = 0.0
                    mins_tramo = 0.0
                    es_ultimo = (i == len(d['paradas']) - 1)
                    espera_real = 0 if es_ultimo else espera
                    
                    if i > 0:
                        secs = d['segmentos'][i-1]['duration'] if i-1 < len(d['segmentos']) else 0
                        meters = d['segmentos'][i-1]['distance'] if i-1 < len(d['segmentos']) else 0
                        dist_tramo = round(meters/1000, 2)
                        mins_tramo = (secs / 60.0) + espera_real
                        t_actual += datetime.timedelta(seconds=secs)
                    else:
                        mins_tramo = espera_real
                    
                    llegada = t_actual
                    if es_ultimo:
                        hora_llegada_final = llegada.strftime("%H:%M")
                        llegada_final_dt = llegada
                        
                    salida = llegada + datetime.timedelta(minutes=espera_real)
                    t_actual = salida
                    
                    dist_acum += dist_tramo
                    mins_acum += mins_tramo
                    
                    rows_excel.append({
                        "Orden": p.get('Orden', i+1),
                        "Día": p['Día'], "Ruta": p['Ruta'],
                        "Departamento": p['Departamento'], "Lugar": p['Lugar'],
                        "Coordenadas": p['Coordenadas'],
                        "Llegada": llegada.strftime("%H:%M"),
                        "Salida": salida.strftime("%H:%M") if not es_ultimo else "-",
                        "Minutos Tramo": round(mins_tramo, 0),
                        "Minutos Acumulados": round(mins_acum, 0),
                        "Km Tramo": dist_tramo,
                        "Km Acumulados": round(dist_acum, 2)
                    })
                
                data_global_detallada.extend(rows_excel)
                
                if llegada_final_dt:
                    minutos_demora_real = int((llegada_final_dt - inicio_dt).total_seconds() / 60)
                else:
                    minutos_demora_real = 0
                    
                waypoints_maps = []
                waypoints_ors = []
                
                for p in d['paradas']:
                    coord_raw = str(p.get('Coordenadas', ''))
                    partes = coord_raw.split(',')
                    if len(partes) >= 2:
                        try:
                            lat = float(partes[0].strip())
                            lon = float(partes[1].strip())
                            
                            # Formato Oficial Google Maps
                            waypoints_maps.append(f"{lat},{lon}")
                            
                            # Formato Oficial ORS Clásico (Siempre Longitud, Latitud)
                            waypoints_ors.append(f"{lon},{lat}")
                        except Exception:
                            pass
                
                # ENLACES CLÁSICOS E INFALIBLES Y LIMPIOS
                enlace_maps = "http://googleusercontent.com/maps.google.com/dir/" + "/".join(waypoints_maps) if waypoints_maps else ""
                enlace_ors = "https://maps.openrouteservice.org/directions?a=" + ",".join(waypoints_ors) + "&b=0&c=0&k1=es-ES&k2=km" if waypoints_ors else ""
                
                data_resumen_general.append({
                    "Día": d['dia'],
                    "Ruta": d['ruta'],
                    "Hs de Inicio": h_inicio.strftime("%H:%M"),
                    "Hs de Finalización": hora_llegada_final,
                    "Minutos de Demora": minutos_demora_real,
                    "Kms Recorridos": round(dist_acum, 2),
                    "Link Google Maps": enlace_maps,
                    "Link ORS": enlace_ors
                })
                
                with st.expander("Ver Cronograma Detallado"):
                    df_view = pd.DataFrame(rows_excel)
                    st.dataframe(df_view[['Orden','Lugar','Llegada','Salida','Minutos Tramo','Minutos Acumulados','Km Acumulados']], use_container_width=True)
                    
                    bio = io.BytesIO()
                    with pd.ExcelWriter(bio, engine='openpyxl') as w:
                        df_view.to_excel(w, index=False, sheet_name=limpiar_nombre_excel(d['id_unico']))
                    st.download_button("📥 Descargar Excel de esta Ruta", bio.getvalue(), f"Ruta_{limpiar_nombre_excel(d['id_unico'])}.xlsx", key=f"dl_{d['id_unico']}")
                st.divider()

    with tab_resumen:
        st.markdown("### Tabla Resumen Operativo")
        st.info("Este es el resumen general para auditar la eficiencia y horarios reales de finalización de todos los autos.")
        
        df_resumen = pd.DataFrame(data_resumen_general)
        
        st.dataframe(
            df_resumen, 
            use_container_width=True,
            column_config={
                "Link Google Maps": st.column_config.LinkColumn("🗺️ G. Maps", display_text="Abrir Maps"),
                "Link ORS": st.column_config.LinkColumn("🧭 ORS", display_text="Abrir ORS")
            }
        )
        
        df_resumen_export = df_resumen.copy()
        if "Link Google Maps" in df_resumen_export.columns:
            df_resumen_export["Link Google Maps"] = df_resumen_export["Link Google Maps"].apply(lambda x: "Abrir Maps" if pd.notna(x) and str(x).startswith("http") else "")
        if "Link ORS" in df_resumen_export.columns:
            df_resumen_export["Link ORS"] = df_resumen_export["Link ORS"].apply(lambda x: "Abrir ORS" if pd.notna(x) and str(x).startswith("http") else "")
            
        bio_resumen = io.BytesIO()
        with pd.ExcelWriter(bio_resumen, engine='openpyxl') as w:
            df_resumen_export.to_excel(w, index=False, sheet_name="Resumen")
            ws = w.sheets["Resumen"]
            font_link = openpyxl.styles.Font(color="0563C1", underline="single")
            
            for r_idx in range(len(df_resumen)):
                row_excel = r_idx + 2
                if 'Link Google Maps' in df_resumen.columns:
                    c_idx = df_resumen.columns.get_loc('Link Google Maps') + 1
                    url_maps = df_resumen.iloc[r_idx]['Link Google Maps']
                    if pd.notna(url_maps) and str(url_maps).startswith("http"):
                        celda = ws.cell(row=row_excel, column=c_idx)
                        celda.hyperlink = str(url_maps)
                        celda.font = font_link
                        
                if 'Link ORS' in df_resumen.columns:
                    c_idx = df_resumen.columns.get_loc('Link ORS') + 1
                    url_ors = df_resumen.iloc[r_idx]['Link ORS']
                    if pd.notna(url_ors) and str(url_ors).startswith("http"):
                        celda = ws.cell(row=row_excel, column=c_idx)
                        celda.hyperlink = str(url_ors)
                        celda.font = font_link
        
        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            st.download_button("📥 DESCARGAR SOLO RESUMEN", bio_resumen.getvalue(), "Resumen_General.xlsx", type="secondary", use_container_width=True)
            
        if data_global_detallada:
            df_glob = pd.DataFrame(data_global_detallada)
            bio_g = io.BytesIO()
            with pd.ExcelWriter(bio_g, engine='openpyxl') as w:
                df_glob.to_excel(w, index=False, sheet_name="Cronograma Detallado")
                df_resumen_export.to_excel(w, index=False, sheet_name="Resumen General") 
                
                ws = w.sheets["Resumen General"]
                font_link = openpyxl.styles.Font(color="0563C1", underline="single")
                for r_idx in range(len(df_resumen)):
                    row_excel = r_idx + 2
                    if 'Link Google Maps' in df_resumen.columns:
                        c_idx = df_resumen.columns.get_loc('Link Google Maps') + 1
                        url_maps = df_resumen.iloc[r_idx]['Link Google Maps']
                        if pd.notna(url_maps) and str(url_maps).startswith("http"):
                            celda = ws.cell(row=row_excel, column=c_idx)
                            celda.hyperlink = str(url_maps)
                            celda.font = font_link
                            
                    if 'Link ORS' in df_resumen.columns:
                        c_idx = df_resumen.columns.get_loc('Link ORS') + 1
                        url_ors = df_resumen.iloc[r_idx]['Link ORS']
                        if pd.notna(url_ors) and str(url_ors).startswith("http"):
                            celda = ws.cell(row=row_excel, column=c_idx)
                            celda.hyperlink = str(url_ors)
                            celda.font = font_link
                            
            with col_btn2:
                st.download_button("📥 DESCARGAR CRONOGRAMA MAESTRO (Completo)", bio_g.getvalue(), "Cronograma_Maestro.xlsx", type="primary", use_container_width=True)
